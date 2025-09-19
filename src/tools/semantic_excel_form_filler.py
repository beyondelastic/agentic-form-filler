"""
Semantic Excel Form Filler Tool - Advanced Excel form filling using semantic extraction and form learning insights.

This tool combines:
1. Form structure anal        # Get all field definitions from the form structure
        all_fields = form_structure.get('all_fields', {})
        if not all_fields:
            print("⚠️ No field definitions found in form structure")
            print(f"🔍 Form structure sections: {form_structure.get('sections', 'No sections')}")
            print(f"🔍 Available form structure keys: {list(form_structure.keys())}")
            return []
        
        print(f"📊 Found {len(all_fields)} field definitions in form structure")rom form_learner and Excel analyzer)
2. Semantic extraction results (from data_extractor) 
3. Intelligent field mapping with LLM assistance
4. Excel-specific optimizations for various field types

Key improvements over basic Excel form filler:
- Leverages comprehensive form structure analysis
- Uses semantic mappings with confidence scoring
- Context-aware field matching
- Support for complex Excel forms with multiple worksheets
- Preserves macros, formatting, and data validation
"""

import os
import json
import shutil
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
from dataclasses import dataclass, asdict
import re

# Excel libraries
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

# Project imports
from src.models import AgentState
from src.llm_client import get_llm_client


@dataclass
class ExcelSemanticMapping:
    """Represents a semantic mapping between extracted data and Excel form field."""
    form_field_name: str
    form_field_id: str  # Cell address like "Sheet1!A1"
    form_field_type: str
    extracted_value: Any
    confidence: float
    extraction_field_id: str  # Original field ID from semantic extraction
    extraction_field_name: str  # Original field name from semantic extraction
    mapping_method: str  # How this mapping was determined
    worksheet_name: str  # Which Excel worksheet
    cell_address: str  # Clean cell reference like "A1"
    context: Optional[str] = None  # Additional context from form structure


@dataclass
class SemanticExcelFillingResult:
    """Enhanced result with semantic mapping information for Excel forms."""
    output_file_path: str
    semantic_mappings: List[ExcelSemanticMapping]
    success: bool
    errors: Optional[List[str]] = None
    total_form_fields: int = 0
    fields_attempted: int = 0
    fields_filled: int = 0
    worksheets_processed: int = 0
    unmapped_extracted_fields: List[str] = None
    unfilled_form_fields: List[str] = None
    mapping_report_path: Optional[str] = None


class SemanticExcelFormFillerTool:
    """
    Advanced Excel form filler that uses semantic extraction and form learning insights.
    
    This tool:
    1. Takes semantic extraction results (field names + values + confidence)
    2. Uses Excel form structure analysis (field positions, types, contexts)
    3. Creates intelligent mappings with LLM assistance
    4. Fills Excel forms with high accuracy using comprehensive context
    5. Preserves Excel formatting, macros, and data validation
    """

    def __init__(self):
        """Initialize the semantic Excel form filler."""
        self.supported_extensions = ['.xlsx', '.xlsm']
        self.llm_client = get_llm_client()
        self.supported_field_types = [
            "text", "number", "date", "email", "phone", 
            "checkbox", "dropdown", "text_area", "currency"
        ]

    async def fill_excel_form_semantically(
        self,
        state: AgentState,
        output_path: str
    ) -> SemanticExcelFillingResult:
        """
        Fill Excel form using semantic extraction results and form learning insights.
        
        Args:
            state: AgentState with form_structure, extracted_data, and other context
            output_path: Where to save the filled Excel form
            
        Returns:
            SemanticExcelFillingResult with detailed mapping information
        """
        
        try:
            print("🧠 Starting semantic Excel form filling...")
            
            # Validate inputs
            if not state.form_template_path or not os.path.exists(state.form_template_path):
                raise FileNotFoundError(f"Excel template not found: {state.form_template_path}")
            
            # Step 1: Create semantic mappings using form learning insights
            semantic_mappings = await self._create_excel_semantic_mappings(state)
            
            print(f"📊 Created {len(semantic_mappings)} semantic mappings for Excel")
            
            # Step 2: Fill the Excel form
            result = await self._fill_excel_with_mappings(
                state.form_template_path, 
                semantic_mappings, 
                output_path,
                state
            )
            
            # Step 3: Save detailed mapping report
            if result.success:
                await self._save_excel_mapping_report(semantic_mappings, state, output_path)
                result.mapping_report_path = self._get_mapping_report_path(output_path)
            
            return result
            
        except Exception as e:
            print(f"❌ Semantic Excel form filling error: {str(e)}")
            return SemanticExcelFillingResult(
                output_file_path="",
                semantic_mappings=[],
                success=False,
                errors=[f"Semantic Excel form filling failed: {str(e)}"]
            )

    async def _create_excel_semantic_mappings(self, state: AgentState) -> List[ExcelSemanticMapping]:
        """
        Create semantic mappings using Excel form structure and extraction results.
        
        This leverages the comprehensive Excel form analysis to create precise mappings.
        """
        mappings = []
        
        # Get form structure from form learning
        form_structure = getattr(state, 'form_structure', None)
        if not form_structure:
            print("⚠️ No form structure available - using basic Excel mapping")
            return await self._create_basic_excel_mappings(state)
        
        # Debug: Check form structure contents
        print(f"🔍 Form structure keys: {list(form_structure.keys()) if form_structure else 'None'}")
        
        # Get extracted data (field_id -> value mapping from semantic extraction)
        extracted_data = state.extracted_data or {}
        
        print(f"🔍 Processing {len(extracted_data)} extracted fields against Excel form structure")
        
        # Get all field definitions from the form structure
        all_fields = form_structure.get('all_fields', {})
        if not all_fields:
            print("⚠️ No field definitions found in form structure")
            return []
        
        print(f"� Found {len(all_fields)} field definitions in form structure")
        
        # Process each field from all_fields (this has the complete field details)
        for field_id, field_def in all_fields.items():
            field_name = field_def.get('name', '')
            field_type = field_def.get('field_type', 'text')
            
            # Excel-specific properties
            worksheet_name = field_def.get('worksheet', 'Sheet1')
            cell_address = field_def.get('cell_address', '')
            context = field_def.get('context', field_name)
            
            if not field_id or not cell_address:
                print(f"⚠️ Skipping field {field_id} - missing cell address")
                continue
            
            # Try to find matching extracted data
            matched_value, confidence, match_method, source_field = await self._find_excel_field_match(
                field_id, field_name, field_type, extracted_data, context
            )
            
            if matched_value is not None:
                mapping = ExcelSemanticMapping(
                    form_field_name=field_name,
                    form_field_id=field_id,
                    form_field_type=field_type,
                    extracted_value=matched_value,
                    confidence=confidence,
                    extraction_field_id=source_field,
                    extraction_field_name=source_field,
                    mapping_method=match_method,
                    worksheet_name=worksheet_name,
                    cell_address=cell_address,
                    context=context
                )
                
                mappings.append(mapping)
                print(f"✅ Mapped {field_name} ({field_id}) → {matched_value} (confidence: {confidence:.1%})")
        
        print(f"📊 Created {len(mappings)} semantic mappings")
        return mappings

    async def _find_excel_field_match(
        self,
        field_id: str,
        field_name: str,
        field_type: str,
        extracted_data: Dict[str, Any],
        context: str
    ) -> Tuple[Any, float, str, str]:
        """
        Find the best match for an Excel field using multiple strategies.
        
        Returns: (matched_value, confidence, match_method, source_field_id)
        """
        
        # Strategy 1: Direct field ID match (highest confidence)
        if field_id in extracted_data:
            return extracted_data[field_id], 1.0, "direct_id_match", field_id
        
        # Strategy 2: Direct field name match
        if field_name in extracted_data:
            return extracted_data[field_name], 0.95, "direct_name_match", field_name
        
        # Strategy 3: Case-insensitive name match
        field_name_lower = field_name.lower()
        for key, value in extracted_data.items():
            if key.lower() == field_name_lower:
                return value, 0.9, "case_insensitive_match", key
        
        # Strategy 4: LLM-based semantic match (context-aware)
        llm_match = await self._find_llm_excel_match(
            field_id, field_name, field_type, context, extracted_data
        )
        
        if llm_match:
            value, confidence, source_field = llm_match
            return value, confidence, "llm_semantic_match", source_field
        
        # Strategy 5: Similarity-based matching
        similarity_match = self._find_similarity_excel_match(field_name, extracted_data)
        if similarity_match:
            value, source_field = similarity_match
            return value, 0.7, "similarity_match", source_field
        
        return None, 0.0, "no_match", ""

    async def _find_llm_excel_match(
        self,
        field_id: str,
        field_name: str,
        field_type: str,
        context: str,
        extracted_data: Dict[str, Any]
    ) -> Optional[Tuple[Any, float, str]]:
        """Use LLM to find semantic matches between Excel form fields and extracted data."""
        
        try:
            # Create context-aware prompt for Excel field matching
            system_prompt = f"""You are an Excel form field matching specialist. Your task is to find the best semantic match between an Excel form field and available extracted data.

CORE MATCHING RULES:
1. Match fields based on semantic meaning and business context
2. Handle different languages (German ↔ English)
3. Consider Excel-specific field types and contexts
4. Prioritize accuracy over partial matches
5. Return confidence score (0.5-1.0) or "NO_MATCH"

EXCEL FIELD CONTEXT:
- Field ID: {field_id}
- Field Name: {field_name}
- Field Type: {field_type}
- Section Context: {context}

MATCHING EXAMPLES:
- Excel field "customer_name" matches extracted "Kundenname" or "Client Name"
- Excel field "invoice_date" matches extracted "Rechnungsdatum" or "Date"
- Excel field "total_amount" matches extracted "Gesamtbetrag" or "Amount"

Response format: "FIELD_NAME|CONFIDENCE" (e.g., "Kundenname|0.85") or "NO_MATCH"."""

            data_keys = list(extracted_data.keys())
            user_message = f"""Excel field to match:
- ID: "{field_id}"
- Name: "{field_name}"
- Type: "{field_type}"
- Context: "{context}"

Available extracted data fields:
{chr(10).join(f"- {key}: {str(extracted_data[key])[:50]}..." for key in data_keys[:20])}

Which extracted field best matches this Excel field? Consider semantic meaning and context."""

            messages = self.llm_client.create_messages(system_prompt, user_message)
            response = await self.llm_client.invoke(messages)
            
            response_content = response.content.strip()
            
            if response_content == "NO_MATCH":
                return None
            
            # Parse response
            if '|' in response_content:
                suggested_field, confidence_str = response_content.split('|', 1)
                try:
                    confidence = float(confidence_str)
                    confidence = max(0.5, min(1.0, confidence))  # Clamp to valid range
                except ValueError:
                    confidence = 0.8  # Default confidence
            else:
                suggested_field = response_content
                confidence = 0.8
            
            # Find the suggested field in our data
            for key, value in extracted_data.items():
                if (key.lower() == suggested_field.lower() or 
                    suggested_field.lower() in key.lower() or
                    key.lower() in suggested_field.lower()):
                    
                    print(f"🧠 LLM matched Excel field '{field_name}' → '{key}' (confidence: {confidence:.1%})")
                    return value, confidence, key
            
            return None
            
        except Exception as e:
            print(f"⚠️ LLM Excel field matching error: {e}")
            return None

    def _find_similarity_excel_match(
        self, 
        field_name: str, 
        extracted_data: Dict[str, Any]
    ) -> Optional[Tuple[Any, str]]:
        """Find match using similarity rules optimized for Excel fields."""
        
        field_name_lower = field_name.lower()
        
        # Clean field name for comparison
        clean_field = self._clean_field_name(field_name_lower)
        
        for key, value in extracted_data.items():
            key_lower = key.lower()
            clean_key = self._clean_field_name(key_lower)
            
            # Various similarity checks
            if (clean_key in clean_field or clean_field in clean_key or
                self._are_excel_fields_similar(clean_field, clean_key)):
                return value, key
        
        return None

    def _clean_field_name(self, field_name: str) -> str:
        """Clean field name for better comparison."""
        # Remove common Excel field prefixes/suffixes and separators
        cleaned = field_name.lower()
        cleaned = re.sub(r'[_\-\s]+', '', cleaned)
        cleaned = re.sub(r'^(txt|field|input|cell)', '', cleaned)
        cleaned = re.sub(r'(txt|field|input|cell)$', '', cleaned)
        return cleaned

    def _are_excel_fields_similar(self, field1: str, field2: str) -> bool:
        """Check if two Excel field names are semantically similar."""
        
        # Common Excel field synonyms
        synonyms = {
            'name': ['namen', 'bezeichnung', 'title'],
            'date': ['datum', 'zeit', 'time'],
            'amount': ['betrag', 'summe', 'total'],
            'address': ['adresse', 'location', 'ort'],
            'phone': ['telefon', 'tel', 'number'],
            'email': ['mail', 'electronic'],
        }
        
        for base_word, synonym_list in synonyms.items():
            if base_word in field1 and any(syn in field2 for syn in synonym_list):
                return True
            if base_word in field2 and any(syn in field1 for syn in synonym_list):
                return True
        
        # Levenshtein-like similarity for short fields
        if len(field1) >= 3 and len(field2) >= 3:
            if self._calculate_similarity(field1, field2) > 0.8:
                return True
        
        return False

    def _calculate_similarity(self, str1: str, str2: str) -> float:
        """Calculate simple similarity score between two strings."""
        if not str1 or not str2:
            return 0.0
        
        # Simple character overlap similarity
        set1 = set(str1.lower())
        set2 = set(str2.lower())
        
        intersection = len(set1.intersection(set2))
        union = len(set1.union(set2))
        
        return intersection / union if union > 0 else 0.0

    async def _create_basic_excel_mappings(self, state: AgentState) -> List[ExcelSemanticMapping]:
        """Create basic mappings when no form structure is available."""
        
        mappings = []
        extracted_data = state.extracted_data or {}
        
        if not extracted_data:
            return mappings
        
        # Try to analyze the Excel template directly
        try:
            from .comprehensive_excel_form_analyzer import ComprehensiveExcelFormAnalyzerTool
            analyzer = ComprehensiveExcelFormAnalyzerTool()
            
            print("📊 Analyzing Excel template for basic mapping...")
            excel_structure = await analyzer.analyze_excel_form_comprehensive(state.form_template_path)
            
            # Use the Excel analyzer results to create mappings
            for section in excel_structure.sections:
                for field in section.fields:
                    # Try to match with extracted data
                    matched_value, confidence, match_method, source_field = await self._find_excel_field_match(
                        field.id, field.name, field.field_type, extracted_data, section.title
                    )
                    
                    if matched_value is not None:
                        mapping = ExcelSemanticMapping(
                            form_field_name=field.name,
                            form_field_id=field.id,
                            form_field_type=field.field_type,
                            extracted_value=matched_value,
                            confidence=confidence,
                            extraction_field_id=source_field,
                            extraction_field_name=source_field,
                            mapping_method=match_method,
                            worksheet_name=field.worksheet or 'Sheet1',
                            cell_address=field.cell_address or field.id,
                            context=section.title
                        )
                        mappings.append(mapping)
                        
        except Exception as e:
            print(f"⚠️ Excel analyzer not available for basic mapping: {e}")
            
            # Fallback: analyze Excel template directly to find proper input cells
            mappings = await self._create_smart_basic_mappings(state.form_template_path, extracted_data)
            
        return mappings
    
    async def _create_smart_basic_mappings(self, template_path: str, extracted_data: dict) -> List[ExcelSemanticMapping]:
        """Create smart basic mappings by analyzing Excel template structure."""
        
        mappings = []
        
        try:
            from openpyxl import load_workbook as openpyxl_load_workbook  
            workbook = openpyxl_load_workbook(template_path, data_only=False)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Find label-input pairs in the sheet
                label_input_pairs = self._find_label_input_pairs(sheet)
                
                print(f"📋 Found {len(label_input_pairs)} potential input fields in {sheet_name}")
                
                # Match extracted data with labels
                for label_cell, input_cell in label_input_pairs:
                    label_text = str(label_cell.value).strip().lower()
                    
                    # Try to match this label with extracted data
                    best_match = None
                    best_confidence = 0
                    best_method = "label_match"
                    
                    for field_name, field_value in extracted_data.items():
                        confidence = self._calculate_label_match_confidence(label_text, field_name)
                        
                        if confidence > best_confidence and confidence > 0.3:  # Minimum threshold
                            best_match = field_value
                            best_confidence = confidence
                    
                    if best_match is not None:
                        mapping = ExcelSemanticMapping(
                            form_field_name=label_text,
                            form_field_id=label_cell.coordinate,
                            form_field_type='text',
                            extracted_value=best_match,
                            confidence=best_confidence,
                            extraction_field_id=best_match,
                            extraction_field_name=best_match,
                            mapping_method=best_method,
                            worksheet_name=sheet_name,
                            cell_address=input_cell.coordinate,
                            context='Smart Basic Mapping'
                        )
                        mappings.append(mapping)
            
            workbook.close()
            
        except Exception as e:
            print(f"⚠️ Error in smart basic mapping: {e}")
            # Ultimate fallback - don't create any mappings if we can't analyze properly
            pass
        
        return mappings
    
    def _find_label_input_pairs(self, sheet):
        """Find label-input cell pairs in the Excel sheet."""
        
        pairs = []
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    # This cell has content (potential label)
                    
                    # Strategy 1: Check right for empty cell (Label | Input pattern)
                    right_cell = sheet.cell(row=row, column=col + 1)
                    if right_cell.value is None:
                        pairs.append((cell, right_cell))
                        continue
                    
                    # Strategy 2: Check below for empty cell (Label above Input pattern)
                    if row < sheet.max_row:
                        below_cell = sheet.cell(row=row + 1, column=col)
                        if below_cell.value is None:
                            pairs.append((cell, below_cell))
                            continue
                    
                    # Strategy 3: For single-column forms, create input cell to the right
                    # This handles forms that are just lists of labels
                    if sheet.max_column == 1:
                        input_cell = sheet.cell(row=row, column=2)  # Column B
                        pairs.append((cell, input_cell))
        
        print(f"🔍 Label-input analysis:")
        print(f"   - Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
        print(f"   - Found {len(pairs)} label-input pairs")
        
        return pairs
    
    def _calculate_label_match_confidence(self, label_text: str, field_name: str) -> float:
        """Calculate confidence of matching a label with a field name."""
        
        label_lower = label_text.lower().strip()
        field_lower = field_name.lower().strip()
        
        # Direct match
        if label_lower == field_lower:
            return 1.0
        
        # Check if field name is contained in label
        if field_lower in label_lower:
            return 0.9
        
        # Check if label is contained in field name
        if label_lower in field_lower:
            return 0.8
        
        # German-English common translations
        translations = {
            'nachname': ['lastname', 'surname', 'family'],
            'vorname': ['firstname', 'given', 'first'],
            'geburtsdatum': ['birthdate', 'birth', 'born'],
            'alter': ['age'],
            'wohnort': ['city', 'place', 'location'],
            'straße': ['street', 'address', 'strasse'],
            'e-mail': ['email', 'mail'],
            'telefon': ['phone', 'tel'],
            'beruf': ['job', 'profession', 'occupation'],
            'firma': ['company', 'employer']
        }
        
        for german, english_list in translations.items():
            if german in label_lower:
                for english in english_list:
                    if english in field_lower:
                        return 0.85
            if german in field_lower:
                for english in english_list:
                    if english in label_lower:
                        return 0.85
        
        # Fuzzy similarity using simple overlap
        common_chars = set(label_lower) & set(field_lower)
        if len(common_chars) >= 3:
            similarity = len(common_chars) / max(len(label_lower), len(field_lower))
            return min(similarity, 0.7)
        
        return 0.0

    async def _fill_excel_with_mappings(
        self,
        template_path: str,
        semantic_mappings: List[ExcelSemanticMapping],
        output_path: str,
        state: AgentState
    ) -> SemanticExcelFillingResult:
        """Fill Excel form using the created semantic mappings."""
        
        try:
            print(f"📝 Filling Excel with {len(semantic_mappings)} semantic mappings")
            
            # Create working copy of the template
            working_copy = self._create_excel_template_copy(template_path, output_path)
            
            # Load workbook (preserve macros and formatting) 
            workbook = None
            filled_mappings = []
            errors = []
            worksheets_processed = set()
            success = False
            
            try:
                # Single workbook load to prevent ZipFile handle conflicts
                # This is the industry-standard solution for openpyxl ZipFile bugs
                workbook = None
                try:
                    # First try: load with VBA support
                    from openpyxl import load_workbook as openpyxl_load_workbook
                    workbook = openpyxl_load_workbook(working_copy, keep_vba=True, data_only=False, read_only=False)
                except Exception as vba_error:
                    print(f"📝 VBA load failed, trying without VBA: {str(vba_error)}")
                    try:
                        # Second try: load without VBA (safer for complex files)  
                        from openpyxl import load_workbook as openpyxl_load_workbook
                        workbook = openpyxl_load_workbook(working_copy, data_only=False, read_only=False)
                    except Exception as load_error:
                        raise RuntimeError(f"Failed to load Excel file: {str(load_error)}")
                
                # Group mappings by worksheet for efficient processing
                mappings_by_worksheet = {}
                for mapping in semantic_mappings:
                    ws_name = mapping.worksheet_name
                    if ws_name not in mappings_by_worksheet:
                        mappings_by_worksheet[ws_name] = []
                    mappings_by_worksheet[ws_name].append(mapping)
                
                # Process each worksheet
                for worksheet_name, worksheet_mappings in mappings_by_worksheet.items():
                    try:
                        # Ensure worksheet exists
                        if worksheet_name not in workbook.sheetnames:
                            if workbook.worksheets:
                                worksheet_name = workbook.worksheets[0].title
                            else:
                                errors.append(f"No worksheets found in workbook")
                                continue
                        
                        worksheet = workbook[worksheet_name]
                        worksheets_processed.add(worksheet_name)
                        
                        print(f"📋 Processing worksheet: {worksheet_name}")
                        
                        # Fill each field in this worksheet
                        for mapping in worksheet_mappings:
                            try:
                                success = await self._fill_excel_cell(
                                    worksheet, mapping, workbook
                                )
                                
                                if success:
                                    filled_mappings.append(mapping)
                                    print(f"✅ Filled {mapping.form_field_name} ({mapping.cell_address}) with '{mapping.extracted_value}' (confidence: {mapping.confidence:.1%})")
                                else:
                                    errors.append(f"Failed to fill field '{mapping.form_field_name}' at {mapping.cell_address}")
                                    
                            except Exception as field_error:
                                error_msg = f"Error filling field '{mapping.form_field_name}': {str(field_error)}"
                                errors.append(error_msg)
                                print(f"⚠️ {error_msg}")
                                
                    except Exception as worksheet_error:
                        error_msg = f"Error processing worksheet '{worksheet_name}': {str(worksheet_error)}"
                        errors.append(error_msg)
                        print(f"❌ {error_msg}")
                
                # Save the filled workbook with robust error handling
                try:
                    # Import required modules
                    import gc
                    import tempfile
                    
                    print("💾 Saving Excel file...")
                    
                    # Method 1: Load and Repack approach (Excel-compatible - same as repacked_test.xlsx)
                    try:
                        # Ensure workbook is properly configured for saving
                        workbook.iso_dates = True  # Ensure proper date handling
                        
                        print("� Using load-and-repack method for Excel compatibility...")
                        
                        # Step 1: Save to temporary Excel file with proper .xlsx extension
                        import tempfile
                        temp_dir = os.path.dirname(output_path)
                        with tempfile.NamedTemporaryFile(suffix='.xlsx', dir=temp_dir, delete=False) as temp_file:
                            temp_path = temp_file.name
                        
                        workbook.save(temp_path)
                        
                        # Step 2: Load and re-save (this is what fixed repacked_test.xlsx)
                        # This process cleans up the internal Excel structure
                        from openpyxl import load_workbook as repack_load_workbook
                        print("📋 Repackaging file for Excel compatibility...")
                        
                        repack_wb = repack_load_workbook(temp_path, data_only=False)
                        repack_wb.save(output_path)  # Save the cleaned version
                        repack_wb.close()
                        
                        # Step 3: Clean up temporary file
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                        
                        # Verify the final file
                        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                            success = len(filled_mappings) > 0
                            print(f"✅ Excel file saved with load-and-repack method: {output_path}")
                        else:
                            raise RuntimeError("Final repacked file is empty or missing")
                            
                    except Exception as repack_save_error:
                        print(f"⚠️ Load-and-repack save failed: {repack_save_error}")
                        print("🔄 Trying alternative save method...")
                        
                        # Method 2: Temporary file approach with better cleanup
                        try:
                            # Create temporary file in same directory as output
                            temp_dir = os.path.dirname(output_path)
                            with tempfile.NamedTemporaryFile(suffix='.xlsx', dir=temp_dir, delete=False) as temp_file:
                                temp_path = temp_file.name
                            
                            # Save to temporary location
                            workbook.save(temp_path)
                            
                            # Verify temporary file
                            if not os.path.exists(temp_path) or os.path.getsize(temp_path) == 0:
                                raise RuntimeError("Temporary Excel file is empty or missing")
                            
                            # Move temp file to final location
                            if os.path.exists(output_path):
                                os.remove(output_path)
                            
                            # Use shutil.move for better cross-platform compatibility
                            import shutil
                            shutil.move(temp_path, output_path)
                            
                            # Final verification
                            if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                                success = len(filled_mappings) > 0
                                print(f"✅ Excel file saved via temp method: {output_path}")
                            else:
                                raise RuntimeError("Final Excel file validation failed")
                                
                        except Exception as temp_save_error:
                            # Clean up temp file if it exists
                            if 'temp_path' in locals() and os.path.exists(temp_path):
                                try:
                                    os.remove(temp_path)
                                except:
                                    pass
                            raise temp_save_error
                    
                except Exception as save_error:
                    # Clean up temp file if it exists
                    temp_output = output_path + '.tmp'
                    if os.path.exists(temp_output):
                        try:
                            os.remove(temp_output)
                        except:
                            pass
                    error_msg = f"Failed to save Excel file: {str(save_error)}"
                    errors.append(error_msg)
                    print(f"❌ {error_msg}")
                    success = False
                
                # Additional file validation after save attempt
                if success and os.path.exists(output_path):
                    try:
                        # Test if the file can be opened by openpyxl
                        from openpyxl import load_workbook as validation_load_workbook
                        test_wb = validation_load_workbook(output_path, data_only=True)
                        test_wb.close()
                        print(f"✅ File validation passed: Excel file is readable")
                    except Exception as validation_error:
                        print(f"⚠️ File validation failed: {validation_error}")
                        error_msg = f"Created Excel file appears corrupted: {validation_error}"
                        errors.append(error_msg)
                        success = False
            
            finally:
                # Simplified but robust workbook cleanup
                if workbook is not None:
                    try:
                        # Close the workbook properly
                        workbook.close()
                    except Exception as close_error:
                        print(f"⚠️ Workbook close warning: {close_error}")
                    finally:
                        workbook = None
                
                # Clean up memory
                import gc
                gc.collect()
            
            # Clean up working copy if different from output
            if working_copy != output_path and os.path.exists(working_copy):
                try:
                    os.remove(working_copy)
                except:
                    pass  # Ignore cleanup errors
            
            return SemanticExcelFillingResult(
                output_file_path=output_path,
                semantic_mappings=filled_mappings,
                success=success,
                errors=errors if errors else None,
                total_form_fields=len(semantic_mappings),
                fields_attempted=len(semantic_mappings),
                fields_filled=len(filled_mappings),
                worksheets_processed=len(worksheets_processed),
                unmapped_extracted_fields=self._find_unmapped_fields(state.extracted_data or {}, semantic_mappings),
                unfilled_form_fields=self._find_unfilled_excel_fields(semantic_mappings, filled_mappings)
            )
            
        except Exception as e:
            return SemanticExcelFillingResult(
                output_file_path="",
                semantic_mappings=[],
                success=False,
                errors=[f"Excel form filling failed: {str(e)}"]
            )

    def _create_excel_template_copy(self, template_path: str, output_path: str) -> str:
        """Create a working copy of the Excel template with validation."""
        
        try:
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Validate the source template first
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Template file not found: {template_path}")
            
            if os.path.getsize(template_path) == 0:
                raise ValueError(f"Template file is empty: {template_path}")
            
            # Test if source template is readable
            try:
                from openpyxl import load_workbook as template_load_workbook
                test_wb = template_load_workbook(template_path, data_only=True)
                test_wb.close()
                print(f"✅ Template validation passed: {template_path}")
            except Exception as template_error:
                print(f"⚠️ Template validation warning: {template_error}")
                # Continue anyway - might still be processable
            
            if template_path != output_path:
                # Copy template to output location with error handling
                try:
                    shutil.copy2(template_path, output_path)
                    print(f"📋 Template copied to: {output_path}")
                except Exception as copy_error:
                    # Try alternative copy method
                    with open(template_path, 'rb') as src:
                        with open(output_path, 'wb') as dst:
                            dst.write(src.read())
                    print(f"📋 Template copied (alternative method) to: {output_path}")
                
                # Verify the copy was successful
                if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
                    raise RuntimeError(f"Failed to create working copy at {output_path}")
                
                return output_path
            else:
                # Create temporary copy
                temp_path = f"{template_path}.tmp"
                shutil.copy2(template_path, temp_path)
                return temp_path
                
        except Exception as e:
            raise RuntimeError(f"Could not create Excel template copy: {str(e)}")

    async def _fill_excel_cell(
        self, 
        worksheet: Worksheet, 
        mapping: ExcelSemanticMapping,
        workbook: Workbook
    ) -> bool:
        """Fill a specific Excel cell with a mapped value."""
        
        try:
            # Extract cell address (remove worksheet name if present)
            cell_address = mapping.cell_address
            if '!' in cell_address:
                # Format is "WorksheetName!CellAddress", extract just the cell part
                cell_address = cell_address.split('!')[-1]
            
            # Get the cell
            cell = worksheet[cell_address]
            
            # Format the value appropriately for Excel
            formatted_value = self._format_value_for_excel_type(
                mapping.extracted_value, 
                mapping.form_field_type,
                mapping.form_field_name
            )
            
            # Set the cell value
            cell.value = formatted_value
            
            return True
            
        except Exception as e:
            print(f"⚠️ Error filling Excel cell {mapping.cell_address}: {str(e)}")
            return False

    def _format_value_for_excel_type(
        self, 
        value: Any, 
        field_type: str, 
        field_name: str
    ) -> Any:
        """Format a value appropriately for Excel based on field type."""
        
        if value is None or str(value).strip() == '':
            return ''
        
        str_value = str(value).strip()
        field_type_lower = field_type.lower()
        
        # Format based on explicit field type
        if field_type_lower == 'number' or field_type_lower == 'currency':
            return self._format_excel_number(str_value)
        elif field_type_lower == 'date':
            return self._format_excel_date(str_value)
        elif field_type_lower == 'phone':
            return self._format_excel_phone(str_value)
        elif field_type_lower == 'email':
            return str_value.lower().strip()
        elif field_type_lower == 'checkbox':
            return self._format_excel_checkbox(str_value)
        else:
            # Infer from field name if type is generic
            field_name_lower = field_name.lower()
            if any(word in field_name_lower for word in ['date', 'birth', 'dob', 'datum']):
                return self._format_excel_date(str_value)
            elif any(word in field_name_lower for word in ['amount', 'price', 'cost', 'betrag', 'preis']):
                return self._format_excel_number(str_value)
            elif any(word in field_name_lower for word in ['phone', 'tel', 'telefon']):
                return self._format_excel_phone(str_value)
            elif any(word in field_name_lower for word in ['email', 'mail']):
                return str_value.lower().strip()
            else:
                return str_value

    def _format_excel_number(self, value: str) -> Any:
        """Format numeric values for Excel."""
        try:
            # Remove common non-numeric characters
            cleaned = re.sub(r'[^\d.,-]', '', value)
            # Handle European decimal notation (comma as decimal separator)
            if ',' in cleaned and '.' in cleaned:
                # Assume comma is thousands separator, dot is decimal
                cleaned = cleaned.replace(',', '')
            elif ',' in cleaned and '.' not in cleaned:
                # Assume comma is decimal separator
                cleaned = cleaned.replace(',', '.')
            
            if '.' in cleaned:
                return float(cleaned)
            else:
                return int(cleaned)
        except (ValueError, TypeError):
            return value  # Return original if can't convert

    def _format_excel_date(self, value: str) -> str:
        """Format date values for Excel."""
        # Excel handles most date formats well, clean up the input
        cleaned = re.sub(r'\s+', ' ', value.strip())
        
        # Common date patterns that Excel recognizes
        date_patterns = [
            r'\d{1,2}[./]\d{1,2}[./]\d{4}',  # DD/MM/YYYY or MM/DD/YYYY
            r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',  # YYYY-MM-DD
            r'\d{1,2}\.\d{1,2}\.\d{4}',      # DD.MM.YYYY (German format)
        ]
        
        return cleaned  # Let Excel handle the parsing

    def _format_excel_phone(self, value: str) -> str:
        """Format phone number for Excel."""
        # Keep the original format but clean it up
        cleaned = re.sub(r'\s+', ' ', value.strip())
        return cleaned

    def _format_excel_checkbox(self, value: str) -> str:
        """Format checkbox/boolean values for Excel."""
        value_lower = str(value).lower().strip()
        
        if value_lower in ['true', 'yes', 'ja', 'x', '✓', 'checked', '1']:
            return 'TRUE'
        elif value_lower in ['false', 'no', 'nein', '', 'unchecked', '0']:
            return 'FALSE'
        else:
            return value  # Return original if unclear

    def _find_unmapped_fields(
        self, 
        extracted_data: Dict[str, Any], 
        mappings: List[ExcelSemanticMapping]
    ) -> List[str]:
        """Find extracted data fields that weren't mapped to any form field."""
        
        mapped_source_fields = {mapping.extraction_field_id for mapping in mappings}
        unmapped = [field_id for field_id in extracted_data.keys() if field_id not in mapped_source_fields]
        
        return unmapped

    def _find_unfilled_excel_fields(
        self, 
        attempted_mappings: List[ExcelSemanticMapping],
        successful_mappings: List[ExcelSemanticMapping]
    ) -> List[str]:
        """Find form fields that couldn't be filled."""
        
        successful_field_ids = {mapping.form_field_id for mapping in successful_mappings}
        unfilled = [
            mapping.form_field_name 
            for mapping in attempted_mappings 
            if mapping.form_field_id not in successful_field_ids
        ]
        
        return unfilled

    async def _save_excel_mapping_report(
        self,
        semantic_mappings: List[ExcelSemanticMapping],
        state: AgentState,
        output_path: str
    ) -> None:
        """Save detailed Excel semantic mapping report."""
        
        try:
            report_path = self._get_mapping_report_path(output_path)
            
            report_data = {
                "generation_info": {
                    "timestamp": datetime.now().isoformat(),
                    "template_path": state.form_template_path,
                    "output_path": output_path,
                    "total_mappings": len(semantic_mappings),
                    "tool_version": "SemanticExcelFormFiller_v1.0"
                },
                "summary_stats": {
                    "high_confidence_mappings": len([m for m in semantic_mappings if m.confidence >= 0.8]),
                    "medium_confidence_mappings": len([m for m in semantic_mappings if 0.5 <= m.confidence < 0.8]),
                    "low_confidence_mappings": len([m for m in semantic_mappings if m.confidence < 0.5]),
                    "worksheets_involved": list(set(m.worksheet_name for m in semantic_mappings)),
                    "mapping_methods": {
                        method: len([m for m in semantic_mappings if m.mapping_method == method])
                        for method in set(m.mapping_method for m in semantic_mappings)
                    }
                },
                "detailed_mappings": [
                    {
                        "form_field_name": mapping.form_field_name,
                        "form_field_id": mapping.form_field_id,
                        "form_field_type": mapping.form_field_type,
                        "worksheet": mapping.worksheet_name,
                        "cell_address": mapping.cell_address,
                        "extracted_value": str(mapping.extracted_value),
                        "confidence": mapping.confidence,
                        "extraction_source": mapping.extraction_field_name,
                        "mapping_method": mapping.mapping_method,
                        "context": mapping.context
                    }
                    for mapping in semantic_mappings
                ]
            }
            
            with open(report_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)
            
            print(f"📊 Excel semantic mapping report saved: {report_path}")
            
        except Exception as e:
            print(f"⚠️ Could not save Excel mapping report: {e}")

    def _get_mapping_report_path(self, output_path: str) -> str:
        """Generate path for the mapping report."""
        base_name = os.path.splitext(os.path.basename(output_path))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_name = f"excel_semantic_mapping_{base_name}_{timestamp}.json"
        
        return os.path.join(
            os.path.dirname(output_path),
            report_name
        )


# Convenience function for integration
async def fill_excel_form_semantically(
    state: AgentState,
    output_path: str
) -> SemanticExcelFillingResult:
    """
    Convenience function to fill Excel forms semantically.
    
    Args:
        state: AgentState with form structure and extracted data
        output_path: Where to save the filled Excel form
        
    Returns:
        SemanticExcelFillingResult with comprehensive information
    """
    
    tool = SemanticExcelFormFillerTool()
    return await tool.fill_excel_form_semantically(state, output_path)
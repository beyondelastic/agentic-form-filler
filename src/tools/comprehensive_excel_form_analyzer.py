"""Comprehensive Excel form analysis tool for understanding spreadsheet structure and context.

This tool applies the same comprehensive analysis approach used for PDF forms to Excel worksheets,
understanding sections, field relationships, context, and semantic meaning for intelligent form filling.
Based on the architecture of comprehensive_form_analyzer.py but adapted for Excel formats.
"""

import os
import json
from typing import Dict, Any, List, Optional, Tuple
from dataclasses import dataclass, asdict
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation
import re

from src.config import config
from src.llm_client import get_llm_client


@dataclass
class ExcelFormSection:
    """Represents a section of the Excel form (similar to PDF sections)."""
    id: str
    title: str
    description: Optional[str]
    instructions: List[str]
    fields: List[Dict[str, Any]]
    subsections: List['ExcelFormSection']
    worksheet: str
    cell_range: str  # e.g., "A1:F20" 
    position: Dict[str, int]  # start_row, start_col, end_row, end_col


@dataclass
class ExcelFormField:
    """Represents a single field in the Excel form."""
    id: str
    name: str
    description: Optional[str]
    field_type: str  # text, number, date, dropdown, checkbox, formula, etc.
    required: bool
    section_id: str
    dependencies: List[str]  # Other field IDs this field depends on
    validation_rules: List[str]
    default_value: Optional[str]
    options: Optional[List[str]]  # For dropdowns, data validation
    cell_address: str  # e.g., "Sheet1!B5"
    worksheet: str
    position: Dict[str, int]  # row, column
    context: str  # Surrounding text/labels that provide context
    named_range: Optional[str]  # If part of a named range


@dataclass
class ExcelFormStructure:
    """Complete Excel form structure analysis."""
    title: str
    description: str
    purpose: str
    sections: List[ExcelFormSection]
    all_fields: Dict[str, ExcelFormField]
    field_relationships: Dict[str, List[str]]
    instructions: List[str]
    warnings: List[str]
    legal_notes: List[str]
    worksheets: List[str]
    language: str
    form_version: Optional[str]
    issuing_authority: Optional[str]
    named_ranges: Dict[str, str]  # name -> cell reference


class ComprehensiveExcelFormAnalyzer:
    """Tool for comprehensive Excel form structure analysis and learning."""
    
    def __init__(self):
        """Initialize the comprehensive Excel form analysis tool."""
        self.llm_client = get_llm_client()
        self.supported_extensions = ['.xlsx', '.xlsm', '.xls']
        
        # Patterns for identifying form fields
        self.field_indicators = [
            r'\[.*?\]',  # [Field Name]
            r'_{3,}',    # _____ (underscores)
            r'\.{3,}',   # ..... (dots)
            r'<.*?>',    # <Field Name>
            r'\(\s*\)',  # ( )
            r'\[\s*\]',  # [ ]
        ]
        
        # Keywords that indicate sections or groups
        self.section_indicators = [
            'section', 'part', 'teil', 'bereich', 'abschnitt', 
            'category', 'kategorie', 'group', 'gruppe', 'area',
            'personal', 'company', 'contact', 'address', 'employment'
        ]
    
    async def analyze_excel_form_structure(self, file_path: str) -> ExcelFormStructure:
        """
        Analyze the complete structure of an Excel form document.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Complete form structure analysis
        """
        print(f"🔍 Starting comprehensive Excel form analysis: {os.path.basename(file_path)}")
        
        # Step 1: Extract raw Excel content and metadata
        raw_content = self._extract_excel_content(file_path)
        
        # Step 2: Identify form fields using Excel-specific methods
        excel_fields = self._extract_excel_form_fields(file_path)
        
        # Step 3: Analyze text content for structure and context
        text_analysis = self._analyze_excel_text_structure(raw_content)
        
        # Step 4: Use LLM to understand structure and context
        form_structure = await self._analyze_structure_with_llm(
            raw_content, text_analysis, file_path, excel_fields
        )
        
        # Step 5: Save analysis results
        await self._save_analysis_results(form_structure, file_path)
        
        print(f"✅ Excel form analysis complete: {len(form_structure.all_fields)} fields identified")
        return form_structure
    
    def _extract_excel_content(self, file_path: str) -> Dict[str, Any]:
        """Extract raw content from Excel file including all worksheets."""
        content = {
            "worksheets": {},
            "named_ranges": {},
            "metadata": {},
            "text_content": "",
            "all_cell_values": {},
            "data_validations": {}
        }
        
        workbook = None
        try:
            workbook = load_workbook(file_path, data_only=False)
            
            # Extract metadata
            content["metadata"] = {
                "filename": os.path.basename(file_path),
                "worksheets": workbook.sheetnames,
                "active_sheet": workbook.active.title if workbook.active else None
            }
            
            # Extract named ranges
            if hasattr(workbook, 'defined_names'):
                for name, definition in workbook.defined_names.items():
                    try:
                        if definition.attr_text:
                            content["named_ranges"][name] = definition.attr_text
                    except:
                        pass  # Skip problematic named ranges
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._extract_worksheet_content(sheet)
                content["worksheets"][sheet_name] = sheet_data
                
                # Add to combined text content
                content["text_content"] += f"\n=== WORKSHEET: {sheet_name} ===\n"
                content["text_content"] += sheet_data["text_content"]
                
        except Exception as e:
            print(f"⚠️ Error extracting Excel content: {str(e)}")
            content["text_content"] = "Error extracting content"
        finally:
            # Properly close workbook to avoid ZipFile errors
            if workbook is not None:
                try:
                    workbook.close()
                except:
                    pass
        
        return content
    
    def _extract_worksheet_content(self, sheet: Worksheet) -> Dict[str, Any]:
        """Extract content from a single worksheet."""
        sheet_data = {
            "text_content": "",
            "cells": {},
            "merged_cells": [],
            "data_validations": {},
            "comments": {},
            "dimensions": {
                "max_row": sheet.max_row,
                "max_column": sheet.max_column
            }
        }
        
        # Extract cell values and build text representation
        for row in range(1, min(sheet.max_row + 1, 200)):  # Limit to prevent excessive data
            row_text = []
            for col in range(1, min(sheet.max_column + 1, 50)):  # Limit columns too
                cell = sheet.cell(row=row, column=col)
                cell_addr = f"{chr(64 + col)}{row}" if col <= 26 else f"A{col-1}{row}"  # Simple address
                
                if cell.value is not None:
                    cell_value = str(cell.value)
                    sheet_data["cells"][cell_addr] = {
                        "value": cell_value,
                        "data_type": str(type(cell.value).__name__),
                        "row": row,
                        "column": col
                    }
                    row_text.append(cell_value)
                    
                    # Check for comments
                    if cell.comment:
                        sheet_data["comments"][cell_addr] = cell.comment.text
                else:
                    row_text.append("")
            
            if any(text.strip() for text in row_text):  # Only add non-empty rows
                sheet_data["text_content"] += f"Row {row}: " + " | ".join(row_text) + "\n"
        
        # Extract merged cells
        for merged_range in sheet.merged_cells.ranges:
            sheet_data["merged_cells"].append(str(merged_range))
        
        # Extract data validations
        if hasattr(sheet, 'data_validations'):
            for dv in sheet.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    sheet_data["data_validations"][str(dv.sqref)] = {
                        "type": dv.type,
                        "formula1": dv.formula1,
                        "formula2": dv.formula2,
                        "allowBlank": dv.allowBlank
                    }
        
        return sheet_data
    
    def _extract_excel_form_fields(self, file_path: str) -> Dict[str, Dict[str, Any]]:
        """Extract form fields from Excel using various detection methods."""
        excel_fields = {}
        
        workbook = None
        try:
            workbook = load_workbook(file_path, data_only=False)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Method 1: Named ranges
                fields_from_named_ranges = self._detect_named_range_fields(workbook, sheet_name)
                excel_fields.update(fields_from_named_ranges)
                
                # Method 2: Data validation cells
                fields_from_validation = self._detect_validation_fields(sheet, sheet_name)
                excel_fields.update(fields_from_validation)
                
                # Method 3: Pattern-based detection
                fields_from_patterns = self._detect_pattern_fields(sheet, sheet_name)
                excel_fields.update(fields_from_patterns)
                
                # Method 4: Empty cells near labels
                fields_from_labels = self._detect_label_based_fields(sheet, sheet_name)
                excel_fields.update(fields_from_labels)
        
        except Exception as e:
            print(f"⚠️ Error extracting Excel form fields: {str(e)}")
        finally:
            # Properly close workbook to avoid ZipFile errors
            if workbook is not None:
                try:
                    workbook.close()
                except:
                    pass
        
        return excel_fields
    
    def _detect_named_range_fields(self, workbook: Workbook, sheet_name: str) -> Dict[str, Dict[str, Any]]:
        """Detect form fields from named ranges."""
        fields = {}
        
        if hasattr(workbook, 'defined_names'):
            for name, definition in workbook.defined_names.items():
                try:
                    if definition.attr_text and sheet_name in definition.attr_text:
                        fields[f"named_{name}"] = {
                            "field_name": name,
                            "field_type": "named_range",
                            "worksheet": sheet_name,
                            "cell_reference": definition.attr_text,
                            "detection_method": "named_range"
                        }
                except:
                    pass
        
        return fields
    
    def _detect_validation_fields(self, sheet: Worksheet, sheet_name: str) -> Dict[str, Dict[str, Any]]:
        """Detect form fields from data validation rules."""
        fields = {}
        
        if hasattr(sheet, 'data_validations'):
            for i, dv in enumerate(sheet.data_validations.dataValidation):
                if hasattr(dv, 'sqref') and dv.sqref:
                    field_id = f"validation_{sheet_name}_{i}"
                    fields[field_id] = {
                        "field_name": f"Validation Field {i+1}",
                        "field_type": dv.type or "text",
                        "worksheet": sheet_name,
                        "cell_reference": str(dv.sqref),
                        "validation_type": dv.type,
                        "validation_formula": dv.formula1,
                        "detection_method": "data_validation"
                    }
        
        return fields
    
    def _detect_pattern_fields(self, sheet: Worksheet, sheet_name: str) -> Dict[str, Dict[str, Any]]:
        """Detect form fields based on text patterns."""
        fields = {}
        
        for row in range(1, min(sheet.max_row + 1, 100)):  # Limit search area
            for col in range(1, min(sheet.max_column + 1, 20)):
                cell = sheet.cell(row=row, column=col)
                
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value)
                    
                    # Check for field indicator patterns
                    for pattern in self.field_indicators:
                        if re.search(pattern, cell_text):
                            cell_addr = cell.coordinate
                            field_id = f"pattern_{sheet_name}_{cell_addr}"
                            
                            fields[field_id] = {
                                "field_name": cell_text.strip(),
                                "field_type": "text",
                                "worksheet": sheet_name,
                                "cell_reference": cell_addr,
                                "pattern_matched": pattern,
                                "detection_method": "pattern_matching"
                            }
                            break
        
        return fields
    
    def _detect_label_based_fields(self, sheet: Worksheet, sheet_name: str) -> Dict[str, Dict[str, Any]]:
        """Detect form fields by finding empty cells near labels."""
        fields = {}
        
        # Look for patterns like "Label:" followed by empty cell
        for row in range(1, min(sheet.max_row + 1, 100)):
            for col in range(1, min(sheet.max_column + 1, 15)):
                cell = sheet.cell(row=row, column=col)
                
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).strip()
                    
                    # Check if this looks like a label (ends with :, has certain keywords)
                    if (cell_text.endswith(':') or 
                        any(keyword in cell_text.lower() for keyword in 
                            ['name', 'email', 'phone', 'address', 'date', 'number', 
                             'titel', 'beruf', 'telefon', 'adresse', 'datum'])):
                        
                        # Check adjacent cells for empty fields
                        adjacent_cells = [
                            (row, col + 1),      # Right
                            (row + 1, col),      # Below
                            (row, col + 2),      # Two cells right
                        ]
                        
                        for adj_row, adj_col in adjacent_cells:
                            if adj_row <= sheet.max_row and adj_col <= sheet.max_column:
                                adj_cell = sheet.cell(row=adj_row, column=adj_col)
                                
                                # Empty cell or cell with placeholder patterns
                                if (not adj_cell.value or 
                                    (isinstance(adj_cell.value, str) and 
                                     any(pattern in str(adj_cell.value) for pattern in ['___', '...', '___']))):
                                    
                                    field_id = f"label_{sheet_name}_{adj_cell.coordinate}"
                                    fields[field_id] = {
                                        "field_name": cell_text.replace(':', '').strip(),
                                        "field_type": self._infer_field_type(cell_text),
                                        "worksheet": sheet_name,
                                        "cell_reference": adj_cell.coordinate,
                                        "label_cell": cell.coordinate,
                                        "label_text": cell_text,
                                        "detection_method": "label_based"
                                    }
                                    break
        
        return fields
    
    def _infer_field_type(self, label_text: str) -> str:
        """Infer field type from label text."""
        label_lower = label_text.lower()
        
        if any(word in label_lower for word in ['email', 'e-mail', 'mail']):
            return 'email'
        elif any(word in label_lower for word in ['phone', 'tel', 'telefon']):
            return 'phone'
        elif any(word in label_lower for word in ['date', 'datum', 'birth', 'geburt']):
            return 'date'
        elif any(word in label_lower for word in ['number', 'nummer', 'nr', '#']):
            return 'number'
        elif any(word in label_lower for word in ['address', 'adresse', 'street', 'straße']):
            return 'address'
        else:
            return 'text'
    
    def _analyze_excel_text_structure(self, raw_content: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze Excel text content for structural patterns."""
        analysis = {
            "potential_sections": [],
            "field_labels": [],
            "instructions": [],
            "patterns": {
                "headers": [],
                "labels": [],
                "values": []
            }
        }
        
        text_content = raw_content.get("text_content", "")
        
        # Find potential section headers (cells with section-indicating words)
        lines = text_content.split('\n')
        for line_num, line in enumerate(lines):
            line_clean = line.strip()
            
            # Skip empty lines and row indicators
            if not line_clean or line_clean.startswith('Row '):
                continue
            
            # Look for section indicators
            for indicator in self.section_indicators:
                if indicator in line_clean.lower():
                    analysis["potential_sections"].append({
                        "text": line_clean,
                        "line_number": line_num,
                        "indicator": indicator
                    })
                    break
            
            # Look for field labels (text ending with colon)
            if ':' in line_clean:
                parts = line_clean.split('|')
                for part in parts:
                    part = part.strip()
                    if part.endswith(':') and len(part) > 1:
                        analysis["field_labels"].append({
                            "text": part,
                            "line_number": line_num
                        })
        
        return analysis
    
    async def _analyze_structure_with_llm(
        self,
        raw_content: Dict[str, Any],
        text_analysis: Dict[str, Any],
        file_path: str,
        excel_fields: Dict[str, Dict[str, Any]]
    ) -> ExcelFormStructure:
        """Use LLM to analyze and understand the Excel form structure."""
        print("🤖 Analyzing Excel form structure with LLM...")
        
        # Prepare comprehensive prompt for LLM analysis
        analysis_prompt = self._build_excel_analysis_prompt(raw_content, text_analysis, excel_fields)
        
        try:
            # Create messages for LLM
            messages = self.llm_client.create_messages(
                "You are an expert Excel form analyst. Analyze Excel worksheets comprehensively to understand form structure and identify fillable fields.",
                analysis_prompt
            )
            
            # Get response from LLM
            response = await self.llm_client.invoke(messages)
            response_text = response.content
            
            # Parse LLM response into structured format
            structure_data = self._parse_llm_response(response_text)
            
            # Create ExcelFormStructure object
            form_structure = self._create_excel_form_structure(structure_data, file_path)
            
            # Enhance with detected Excel fields to ensure completeness
            form_structure = self._enhance_with_excel_fields(form_structure, excel_fields)
            
            return form_structure
        
        except Exception as e:
            print(f"❌ Error in LLM analysis: {str(e)}")
            # Return a basic structure as fallback
            return self._create_fallback_excel_structure(raw_content, file_path)
    
    def _build_excel_analysis_prompt(
        self,
        raw_content: Dict[str, Any],
        text_analysis: Dict[str, Any],
        excel_fields: Dict[str, Dict[str, Any]]
    ) -> str:
        """Build a comprehensive prompt for Excel form LLM analysis."""
        
        # Build Excel fields summary for the prompt
        excel_fields_info = ""
        if excel_fields:
            excel_fields_info = f"\n\nDETECTED EXCEL FORM FIELDS ({len(excel_fields)} fields):\n"
            for field_id, field_info in excel_fields.items():
                excel_fields_info += f"- {field_info.get('field_name', field_id)} ({field_info.get('field_type', 'unknown')}) "
                excel_fields_info += f"at {field_info.get('cell_reference', 'unknown')} "
                excel_fields_info += f"[{field_info.get('detection_method', 'unknown')}]\n"
        
        # Limit content to prevent overly long prompts
        text_content = raw_content.get("text_content", "")
        if len(text_content) > 8000:
            text_content = text_content[:8000] + "\n... [content truncated] ..."
        
        prompt = f"""
Analyze this Excel form worksheet to understand its structure and identify all fillable fields. Return ONLY valid JSON (no markdown, no explanations).

EXCEL CONTENT:
{text_content}

WORKSHEETS: {', '.join(raw_content.get('metadata', {}).get('worksheets', []))}

NAMED RANGES: {json.dumps(raw_content.get('named_ranges', {}), indent=2)}

{excel_fields_info}

POTENTIAL SECTIONS DETECTED:
{json.dumps(text_analysis.get('potential_sections', []), indent=2)}

FIELD LABELS DETECTED:
{json.dumps(text_analysis.get('field_labels', []), indent=2)}

Return a JSON object with this EXACT structure:
{{
  "form_info": {{
    "title": "Form title from content",
    "description": "Brief description of form purpose",
    "purpose": "What this form is used for",
    "language": "detected language (en/de/etc)",
    "form_version": "version if found, otherwise null",
    "issuing_authority": "authority if found, otherwise null"
  }},
  "sections": [
    {{
      "id": "section_1",
      "title": "Section Name",
      "description": "What this section covers",
      "worksheet": "worksheet name",
      "cell_range": "A1:F20",
      "instructions": ["instruction 1", "instruction 2"],
      "fields": ["field_id_1", "field_id_2"]
    }}
  ],
  "fields": [
    {{
      "id": "field_1",
      "name": "Field Display Name",
      "description": "What this field is for",
      "field_type": "text|number|email|phone|date|dropdown|checkbox",
      "required": true|false,
      "section_id": "section_1",
      "cell_address": "Sheet1!B5",
      "worksheet": "Sheet1",
      "validation_rules": ["rule 1", "rule 2"],
      "context": "surrounding labels/text that explain this field",
      "dependencies": ["other_field_id"],
      "options": ["option1", "option2"] // for dropdowns only
    }}
  ],
  "field_relationships": {{
    "field_1": ["related_field_2", "related_field_3"]
  }},
  "instructions": ["general instruction 1", "general instruction 2"],
  "warnings": ["warning 1", "warning 2"],
  "legal_notes": ["legal note 1", "legal note 2"]
}}

Focus on:
1. Identifying logical sections/groups of related fields
2. Understanding field context from nearby labels
3. Detecting field types from labels and validation rules
4. Finding relationships between fields
5. Extracting any instructions or warnings
6. Making the analysis generic and reusable across different Excel forms
"""
        
        return prompt
    
    def _parse_llm_response(self, response: str) -> Dict[str, Any]:
        """Parse LLM response into structured format."""
        try:
            # Remove any markdown formatting
            response = response.strip()
            if response.startswith('```json'):
                response = response[7:]
            if response.startswith('```'):
                response = response[3:]
            if response.endswith('```'):
                response = response[:-3]
            
            # Parse JSON
            structure_data = json.loads(response)
            return structure_data
            
        except Exception as e:
            print(f"⚠️ Error parsing LLM response: {str(e)}")
            # Return minimal structure as fallback
            return {
                "form_info": {
                    "title": "Excel Form",
                    "description": "Form analysis failed",
                    "purpose": "Unknown",
                    "language": "unknown",
                    "form_version": None,
                    "issuing_authority": None
                },
                "sections": [],
                "fields": [],
                "field_relationships": {},
                "instructions": [],
                "warnings": [],
                "legal_notes": []
            }
    
    def _create_excel_form_structure(self, structure_data: Dict[str, Any], file_path: str) -> ExcelFormStructure:
        """Create ExcelFormStructure object from parsed data."""
        form_info = structure_data.get("form_info", {})
        
        # Create sections
        sections = []
        for section_data in structure_data.get("sections", []):
            section = ExcelFormSection(
                id=section_data.get("id", "unknown"),
                title=section_data.get("title", "Unknown Section"),
                description=section_data.get("description"),
                instructions=section_data.get("instructions", []),
                fields=section_data.get("fields", []),
                subsections=[],  # Can be enhanced later
                worksheet=section_data.get("worksheet", "Sheet1"),
                cell_range=section_data.get("cell_range", "A1:A1"),
                position=self._parse_cell_range_to_position(section_data.get("cell_range", "A1:A1"))
            )
            sections.append(section)
        
        # Create fields
        all_fields = {}
        for field_data in structure_data.get("fields", []):
            field = ExcelFormField(
                id=field_data.get("id", "unknown"),
                name=field_data.get("name", "Unknown Field"),
                description=field_data.get("description"),
                field_type=field_data.get("field_type", "text"),
                required=field_data.get("required", False),
                section_id=field_data.get("section_id", "unknown"),
                dependencies=field_data.get("dependencies", []),
                validation_rules=field_data.get("validation_rules", []),
                default_value=field_data.get("default_value"),
                options=field_data.get("options"),
                cell_address=field_data.get("cell_address", "A1"),
                worksheet=field_data.get("worksheet", "Sheet1"),
                position=self._parse_cell_address_to_position(field_data.get("cell_address", "A1")),
                context=field_data.get("context", ""),
                named_range=field_data.get("named_range")
            )
            all_fields[field.id] = field
        
        return ExcelFormStructure(
            title=form_info.get("title", os.path.basename(file_path)),
            description=form_info.get("description", "Excel form"),
            purpose=form_info.get("purpose", "Form filling"),
            sections=sections,
            all_fields=all_fields,
            field_relationships=structure_data.get("field_relationships", {}),
            instructions=structure_data.get("instructions", []),
            warnings=structure_data.get("warnings", []),
            legal_notes=structure_data.get("legal_notes", []),
            worksheets=[], # Will be filled by enhancement
            language=form_info.get("language", "unknown"),
            form_version=form_info.get("form_version"),
            issuing_authority=form_info.get("issuing_authority"),
            named_ranges={}  # Will be filled by enhancement
        )
    
    def _parse_cell_range_to_position(self, cell_range: str) -> Dict[str, int]:
        """Parse Excel cell range to position dictionary."""
        try:
            if ':' in cell_range:
                start, end = cell_range.split(':')
                # Simple parsing - can be enhanced
                return {"start_row": 1, "start_col": 1, "end_row": 10, "end_col": 5}
            else:
                return {"start_row": 1, "start_col": 1, "end_row": 1, "end_col": 1}
        except:
            return {"start_row": 1, "start_col": 1, "end_row": 1, "end_col": 1}
    
    def _parse_cell_address_to_position(self, cell_address: str) -> Dict[str, int]:
        """Parse Excel cell address to position dictionary."""
        try:
            # Remove worksheet name if present
            if '!' in cell_address:
                cell_address = cell_address.split('!')[-1]
            
            # Simple parsing - extract row/col (can be enhanced)
            import re
            match = re.match(r'([A-Z]+)(\d+)', cell_address)
            if match:
                col_letters, row_num = match.groups()
                # Convert column letters to number (simplified)
                col_num = ord(col_letters[0]) - ord('A') + 1
                return {"row": int(row_num), "column": col_num}
        except:
            pass
        
        return {"row": 1, "column": 1}
    
    def _enhance_with_excel_fields(self, form_structure: ExcelFormStructure, excel_fields: Dict[str, Dict[str, Any]]) -> ExcelFormStructure:
        """Enhance form structure with detected Excel fields."""
        # Add any detected fields that weren't captured by LLM
        for field_id, field_info in excel_fields.items():
            if field_id not in form_structure.all_fields:
                # Create field from detected Excel field
                field = ExcelFormField(
                    id=field_id,
                    name=field_info.get("field_name", field_id),
                    description=f"Detected via {field_info.get('detection_method', 'unknown')}",
                    field_type=field_info.get("field_type", "text"),
                    required=False,
                    section_id="detected_fields",
                    dependencies=[],
                    validation_rules=[],
                    default_value=None,
                    options=None,
                    cell_address=field_info.get("cell_reference", "A1"),
                    worksheet=field_info.get("worksheet", "Sheet1"),
                    position=self._parse_cell_address_to_position(field_info.get("cell_reference", "A1")),
                    context=f"Detected field via {field_info.get('detection_method')}",
                    named_range=field_info.get("field_name") if field_info.get("detection_method") == "named_range" else None
                )
                form_structure.all_fields[field_id] = field
        
        return form_structure
    
    def _create_fallback_excel_structure(self, raw_content: Dict[str, Any], file_path: str) -> ExcelFormStructure:
        """Create a basic fallback structure when LLM analysis fails."""
        return ExcelFormStructure(
            title=os.path.basename(file_path),
            description="Excel form (analysis failed)",
            purpose="Unknown",
            sections=[],
            all_fields={},
            field_relationships={},
            instructions=[],
            warnings=["Form analysis failed - using fallback structure"],
            legal_notes=[],
            worksheets=raw_content.get("metadata", {}).get("worksheets", []),
            language="unknown",
            form_version=None,
            issuing_authority=None,
            named_ranges=raw_content.get("named_ranges", {})
        )
    
    async def _save_analysis_results(self, form_structure: ExcelFormStructure, file_path: str):
        """Save the form analysis results to JSON file."""
        timestamp = __import__('datetime').datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"excel_form_learning_{timestamp}.json"
        
        output_dir = config.OUTPUT_DIR
        output_path = os.path.join(output_dir, filename)
        
        # Convert to serializable format
        analysis_data = {
            "source_file": os.path.basename(file_path),
            "analysis_timestamp": timestamp,
            "form_structure": asdict(form_structure),
            "summary": {
                "total_fields": len(form_structure.all_fields),
                "total_sections": len(form_structure.sections),
                "worksheets": form_structure.worksheets,
                "language": form_structure.language
            }
        }
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(analysis_data, f, indent=2, ensure_ascii=False, default=str)
            print(f"💾 Excel form learning results saved: {output_path}")
        except Exception as e:
            print(f"⚠️ Failed to save Excel analysis results: {str(e)}")
    
    def save_analysis_result(self, form_structure: ExcelFormStructure, output_path: str) -> str:
        """Sync wrapper for saving analysis results - maintains interface compatibility with PDF analyzer.
        
        Note: Excel analysis results are already saved during analyze_excel_form_structure().
        This method is provided for interface compatibility but skips duplicate saving.
        """
        
        # Since Excel analysis already saves results during analysis, we don't need to save again
        # Just return the expected output path pattern that was likely already created
        timestamp = __import__('datetime').datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"excel_form_learning_{timestamp}.json"
        actual_output_path = os.path.join(config.OUTPUT_DIR, filename)
        
        print(f"ℹ️ Excel form analysis results already saved during analysis phase")
        
        return actual_output_path
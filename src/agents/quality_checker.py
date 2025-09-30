"""Quality Checker Agent that validates filled forms against reference patterns.

This agent:
1. Analyzes reference forms to learn correct field patterns
2. Compares filled forms against these patterns  
3. Identifies semantic mismatches and contextual errors
4. Provides specific feedback for corrections
5. Supports iterative improvement cycles

Designed to be generic and work with both Excel and PDF forms.
"""

import os
import json
from typing import Dict, Any, Optional, List
from datetime import datetime

from src.models import (
    AgentState, AgentType, QualityAssessmentResult, 
    QualityIssue, ReferenceFieldPattern
)
from src.llm_client import get_llm_client


class QualityCheckerAgent:
    """
    Quality Checker Agent that validates form filling quality using reference patterns.
    
    This agent provides a generic approach to quality validation by:
    1. Learning patterns from reference forms (any format)
    2. Assessing filled forms against these patterns
    3. Generating targeted feedback for improvements
    4. Supporting iterative correction cycles
    """
    
    def __init__(self):
        self.agent_type = AgentType.QUALITY_CHECKER
        self.llm_client = get_llm_client()
        self.reference_cache = {}  # Cache analyzed reference forms
    
    async def process(self, state: AgentState) -> AgentState:
        """Process quality checking of filled forms."""
        print(f"\n🔍 Quality Checker Agent Processing")
        
        try:
            # Check if we have a filled form to assess
            if not state.filled_form_path or not os.path.exists(state.filled_form_path):
                return self._handle_quality_error(state, "No filled form available for quality assessment")
            
            # Step 1: Analyze reference form if available
            reference_patterns = await self._get_reference_patterns(state)
            
            # Step 2: Assess the filled form quality
            quality_result = await self._assess_form_quality(state, reference_patterns)
            
            # Step 3: Update state with quality results
            state.quality_assessment = quality_result
            state.reference_patterns = reference_patterns
            
            # Step 4: Determine next steps based on quality assessment
            return self._determine_next_steps(state, quality_result)
            
        except Exception as e:
            return self._handle_quality_error(state, f"Quality checking error: {str(e)}")
    
    async def _get_reference_patterns(self, state: AgentState) -> Optional[Dict[str, ReferenceFieldPattern]]:
        """Get or analyze reference form patterns."""
        
        # Check if reference form is specified
        if not state.reference_form_path:
            print("📝 No reference form specified - using basic quality checks")
            return None
        
        if not os.path.exists(state.reference_form_path):
            print(f"⚠️  Reference form not found: {state.reference_form_path}")
            return None
        
        # Check cache first
        cache_key = f"{state.reference_form_path}_{os.path.getmtime(state.reference_form_path)}"
        if cache_key in self.reference_cache:
            print("📚 Using cached reference patterns")
            return self.reference_cache[cache_key]
        
        print(f"📖 Analyzing reference form: {os.path.basename(state.reference_form_path)}")
        
        # Analyze reference form based on file type
        file_extension = os.path.splitext(state.reference_form_path)[1].lower()
        
        if file_extension in ['.xlsx', '.xlsm']:
            patterns = await self._analyze_excel_reference(state.reference_form_path, state)
        elif file_extension == '.pdf':
            patterns = await self._analyze_pdf_reference(state.reference_form_path, state)
        else:
            print(f"⚠️  Unsupported reference form format: {file_extension}")
            return None
        
        # Cache the results
        self.reference_cache[cache_key] = patterns
        return patterns
    
    async def _analyze_excel_reference(self, reference_path: str, state: AgentState) -> Dict[str, ReferenceFieldPattern]:
        """Analyze Excel reference form to learn field patterns."""
        
        try:
            import openpyxl
            
            # Load reference form with values
            wb = openpyxl.load_workbook(reference_path, data_only=True)
            patterns = {}
            
            # For now, focus on the first worksheet (can be made generic later)
            sheet_name = wb.sheetnames[0]
            sheet = wb[sheet_name]
            
            print(f"   📊 Analyzing worksheet: {sheet_name}")
            
            # Extract field patterns by analyzing the structure
            # This is a basic implementation - can be enhanced based on form structure
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and ':' not in str(cell.value):
                        # Look for field labels in column A and values in column B
                        if cell.column == 1 and cell.value.strip():  # Column A (labels)
                            field_label = str(cell.value).strip()
                            value_cell = sheet.cell(row=cell.row, column=2)  # Column B
                            
                            if value_cell.value:
                                field_id = f"field_{cell.row-1}"  # Generate field ID based on position
                                pattern = self._create_field_pattern(
                                    field_id, field_label, value_cell.value
                                )
                                patterns[field_id] = pattern
                                print(f"   📌 Pattern learned: {field_label} -> {self._truncate_value(value_cell.value)}")
            
            print(f"   ✅ Learned {len(patterns)} field patterns from reference")
            return patterns
            
        except Exception as e:
            print(f"   ❌ Error analyzing Excel reference: {str(e)}")
            return {}
    
    async def _analyze_pdf_reference(self, reference_path: str, state: AgentState) -> Dict[str, ReferenceFieldPattern]:
        """Analyze PDF reference form to learn field patterns."""
        
        try:
            from src.tools.comprehensive_form_analyzer import ComprehensiveFormAnalysisTool
            
            print("   📄 Analyzing PDF reference form...")
            
            # Use the comprehensive form analyzer to extract PDF form structure
            analyzer = ComprehensiveFormAnalysisTool()
            pdf_fields = analyzer._extract_pdf_form_fields(reference_path)
            
            # Convert PDF fields to reference patterns
            patterns = {}
            for field_name, field_data in pdf_fields.items():
                if field_data.get('current_value'):  # Only analyze fields with reference values
                    pattern = self._create_field_pattern(
                        field_id=field_name,
                        field_name=field_name, 
                        reference_value=field_data['current_value']
                    )
                    patterns[field_name] = pattern
            
            print(f"   📋 Created {len(patterns)} reference patterns from PDF form")
            return patterns
            
        except Exception as e:
            print(f"   ❌ Error analyzing PDF reference: {str(e)}")
            return {}
    
    def _create_field_pattern(self, field_id: str, field_name: str, reference_value: Any) -> ReferenceFieldPattern:
        """Create a field pattern from reference data."""
        
        # Determine semantic category based on field name and value
        semantic_category = self._determine_semantic_category(field_name, reference_value)
        
        # Determine field type
        field_type = self._determine_field_type(reference_value)
        
        return ReferenceFieldPattern(
            field_id=field_id,
            field_name=field_name,
            field_type=field_type,
            semantic_category=semantic_category,
            value_pattern=self._extract_value_pattern(reference_value),
            context_clues=[field_name.lower()],
            example_values=[str(reference_value)]
        )
    
    def _determine_semantic_category(self, field_name: str, value: Any) -> str:
        """Determine the semantic category of a field."""
        
        field_lower = field_name.lower()
        value_str = str(value).lower() if value else ""
        
        # Date fields
        if any(keyword in field_lower for keyword in ['datum', 'date']):
            if any(keyword in field_lower for keyword in ['eingang', 'received', 'submit', 'application']):
                return "document_date"
            elif any(keyword in field_lower for keyword in ['geburt', 'birth', 'born']):
                return "personal_date"
            else:
                return "date_general"
        
        # Name fields
        if any(keyword in field_lower for keyword in ['name', 'vor', 'nach', 'first', 'last']):
            return "personal_name"
        
        # Contact fields
        if any(keyword in field_lower for keyword in ['mail', 'email', 'telefon', 'phone']):
            return "contact_info"
        
        # Address fields
        if any(keyword in field_lower for keyword in ['straße', 'street', 'ort', 'city', 'address']):
            return "address_info"
        
        # Age/numeric fields
        if any(keyword in field_lower for keyword in ['alter', 'age', 'note', 'grade']):
            return "numeric_value"
        
        return "general_field"
    
    def _determine_field_type(self, value: Any) -> str:
        """Determine the data type of a field."""
        
        if isinstance(value, (int, float)):
            return "number"
        elif isinstance(value, str):
            # Check if it looks like a date
            if self._looks_like_date(value):
                return "date"
            elif '@' in value:
                return "email"
            else:
                return "text"
        else:
            return "text"
    
    def _looks_like_date(self, value: str) -> bool:
        """Check if a string looks like a date."""
        import re
        date_patterns = [
            r'\d{1,2}\.\d{1,2}\.\d{4}',    # DD.MM.YYYY
            r'\d{1,2}\.\d{1,2}\.\d{2}',    # DD.MM.YY (2-digit year)
            r'\d{4}-\d{1,2}-\d{1,2}',     # YYYY-MM-DD
            r'\d{1,2}/\d{1,2}/\d{4}',     # MM/DD/YYYY
            r'\d{1,2}/\d{1,2}/\d{2}',     # MM/DD/YY
        ]
        return any(re.match(pattern, value.strip()) for pattern in date_patterns)
    
    def _extract_value_pattern(self, value: Any) -> Optional[str]:
        """Extract a pattern description from a reference value."""
        
        if self._looks_like_date(str(value)):
            return "date_format"
        elif isinstance(value, (int, float)):
            return "numeric_value"
        elif '@' in str(value):
            return "email_format"
        else:
            return "text_value"
    
    async def _assess_form_quality(self, state: AgentState, reference_patterns: Optional[Dict[str, ReferenceFieldPattern]]) -> QualityAssessmentResult:
        """Assess the quality of the filled form."""
        
        print("🔍 Assessing form quality...")
        
        issues = []
        total_checks = 0
        passed_checks = 0
        
        # Basic quality checks (without reference)
        basic_issues, basic_total, basic_passed = await self._perform_basic_quality_checks(state)
        issues.extend(basic_issues)
        total_checks += basic_total
        passed_checks += basic_passed
        
        # Reference-based quality checks (if reference patterns available)
        if reference_patterns:
            ref_issues, ref_total, ref_passed = await self._perform_reference_quality_checks(state, reference_patterns)
            issues.extend(ref_issues)
            total_checks += ref_total
            passed_checks += ref_passed
        
        # Calculate overall quality score
        quality_score = (passed_checks / total_checks) if total_checks > 0 else 0.0
        
        print(f"   📊 Quality assessment: {passed_checks}/{total_checks} checks passed ({quality_score:.1%})")
        print(f"   🚨 Found {len(issues)} issues")
        
        # Save detailed quality report for debugging
        quality_result = QualityAssessmentResult(
            overall_quality_score=quality_score,
            issues_found=issues,
            passed_checks=passed_checks,
            total_checks=total_checks,
            reference_form_used=state.reference_form_path,
            assessment_timestamp=datetime.now().isoformat(),
            requires_correction=len([i for i in issues if i.severity in ["medium", "high", "critical"]]) > 0
        )
        
        # Save detailed JSON report
        await self._save_quality_report(quality_result, state)
        
        return quality_result
    
    async def _perform_basic_quality_checks(self, state: AgentState) -> tuple[List[QualityIssue], int, int]:
        """Perform basic quality checks without reference patterns."""
        
        print("   ✓ Performing basic quality checks (no reference form provided)...")
        print("   ℹ️  Note: Limited validation without reference form - format and basic semantic checks only")
        
        issues = []
        total_checks = 0
        passed_checks = 0
        
        # Get current filled values
        current_values = await self._extract_current_form_values(state)
        
        for field_id, value in current_values.items():
            total_checks += 1
            
            # Basic format validations
            basic_issue = self._check_basic_format(field_id, value)
            if basic_issue:
                issues.append(basic_issue)
            else:
                passed_checks += 1
                
            # Enhanced semantic checks (even without reference form)
            enhanced_issue = self._check_enhanced_basic_semantics(field_id, value, current_values)
            if enhanced_issue:
                issues.append(enhanced_issue)
                total_checks += 1  # Add to total since this is an additional check
        
        # Check for completely empty form
        if not current_values:
            issues.append(QualityIssue(
                field_id="form_general",
                field_name="Form Completeness",
                issue_type="empty_form",
                current_value=None,
                confidence=1.0,
                suggestion="Form appears to be empty or unreadable",
                severity="critical"
            ))
            total_checks = 1
        
        if total_checks > 0:
            print(f"   📊 Basic checks: {passed_checks}/{total_checks} passed")
            if len(issues) == 0 and not state.reference_form_path:
                print(f"   ⚠️  Note: 100% score with basic checks only - provide reference form for comprehensive validation")
        else:
            print("   ⚠️  No fields found for quality checking")
            
        return issues, total_checks, passed_checks
    
    def _check_basic_format(self, field_id: str, value: Any) -> Optional[QualityIssue]:
        """Perform basic format validation on a field value."""
        
        if value is None or str(value).strip() == "":
            return None  # Empty fields are handled elsewhere
        
        value_str = str(value).strip()
        
        # Check for obvious data issues
        if len(value_str) > 1000:  # Extremely long values
            return QualityIssue(
                field_id=field_id,
                field_name=f"Field {field_id}",
                issue_type="format_error",
                current_value=self._truncate_value(value),
                confidence=0.8,
                suggestion="Field value is extremely long and may contain errors",
                severity="medium"
            )
        
        # Check for unusual characters that might indicate extraction errors
        if any(char in value_str for char in ['§', '©', '®', '™']) and len(value_str) < 50:
            return QualityIssue(
                field_id=field_id,
                field_name=f"Field {field_id}",
                issue_type="format_error",
                current_value=value,
                confidence=0.7,
                suggestion="Field contains unusual characters that might indicate extraction errors",
                severity="low"
            )
        
        return None
    
    def _check_enhanced_basic_semantics(self, field_id: str, value: Any, all_values: Dict[str, Any]) -> Optional[QualityIssue]:
        """Perform enhanced semantic checks without requiring a reference form."""
        
        if value is None or str(value).strip() == "":
            return None
        
        value_str = str(value).strip()
        field_id_lower = field_id.lower()
        
        # Date field semantic validation
        if self._looks_like_date(value_str):
            # Check for potential date field mismatches
            if any(keyword in field_id_lower for keyword in ['name', 'namen', 'company', 'firma', 'address', 'strasse']):
                return QualityIssue(
                    field_id=field_id,
                    field_name=f"Field {field_id}",
                    issue_type="semantic_mismatch",
                    current_value=value,
                    confidence=0.8,
                    suggestion=f"Field appears to be for name/address but contains a date: '{value_str}'",
                    severity="medium"
                )
        
        # Name field validation
        elif any(keyword in field_id_lower for keyword in ['name', 'namen', 'vorname', 'nachname']):
            if self._looks_like_date(value_str):
                return QualityIssue(
                    field_id=field_id,
                    field_name=f"Field {field_id}",
                    issue_type="semantic_mismatch",
                    current_value=value,
                    confidence=0.9,
                    suggestion=f"Name field contains what appears to be a date: '{value_str}'",
                    severity="high"
                )
            elif len(value_str) > 100:
                return QualityIssue(
                    field_id=field_id,
                    field_name=f"Field {field_id}",
                    issue_type="format_error",
                    current_value=self._truncate_value(value),
                    confidence=0.7,
                    suggestion="Name field is unusually long and may contain extracted text fragments",
                    severity="medium"
                )
        
        # Phone/email validation
        elif any(keyword in field_id_lower for keyword in ['phone', 'telefon', 'email', 'mail']):
            if 'email' in field_id_lower or 'mail' in field_id_lower:
                if '@' not in value_str:
                    return QualityIssue(
                        field_id=field_id,
                        field_name=f"Field {field_id}",
                        issue_type="format_error",
                        current_value=value,
                        confidence=0.8,
                        suggestion=f"Email field does not contain '@' symbol: '{value_str}'",
                        severity="medium"
                    )
            elif 'phone' in field_id_lower or 'telefon' in field_id_lower:
                if not any(char.isdigit() for char in value_str):
                    return QualityIssue(
                        field_id=field_id,
                        field_name=f"Field {field_id}",
                        issue_type="format_error",
                        current_value=value,
                        confidence=0.8,
                        suggestion=f"Phone field does not contain any digits: '{value_str}'",
                        severity="medium"
                    )
        
        return None
    
    async def _perform_reference_quality_checks(self, state: AgentState, reference_patterns: Dict[str, ReferenceFieldPattern]) -> tuple[List[QualityIssue], int, int]:
        """Perform reference-based quality checks using learned patterns."""
        
        print("   🔍 Performing reference-based quality checks...")
        
        issues = []
        total_checks = 0
        passed_checks = 0
        
        # Get current filled values from the form
        current_values = await self._extract_current_form_values(state)
        
        # Compare each field against reference patterns
        for field_id, pattern in reference_patterns.items():
            total_checks += 1
            
            if field_id not in current_values:
                # Field not filled - create issue
                issues.append(QualityIssue(
                    field_id=field_id,
                    field_name=pattern.field_name,
                    issue_type="missing_field",
                    current_value=None,
                    confidence=0.9,
                    suggestion=f"Field '{pattern.field_name}' should be filled based on reference pattern",
                    severity="medium"
                ))
                continue
            
            current_value = current_values[field_id]
            
            # Perform semantic validation
            semantic_issue = self._check_semantic_consistency(
                pattern, current_value, current_values, state
            )
            if semantic_issue:
                issues.append(semantic_issue)
            else:
                passed_checks += 1
                
            # Perform format validation
            format_issue = self._check_format_consistency(pattern, current_value)
            if format_issue:
                issues.append(format_issue)
            
            # Perform contextual validation using all available knowledge
            contextual_issue = await self._check_contextual_consistency(
                pattern, current_value, current_values, reference_patterns, state
            )
            if contextual_issue:
                issues.append(contextual_issue)
        
        print(f"   📊 Reference checks: {passed_checks}/{total_checks} passed, {len(issues)} issues found")
        
        return issues, total_checks, passed_checks
    
    def _determine_next_steps(self, state: AgentState, quality_result: QualityAssessmentResult) -> AgentState:
        """Determine next steps based on quality assessment."""
        
        if quality_result.requires_correction and state.quality_iteration_count < state.max_quality_iterations:
            # Issues found - route back for correction
            state.quality_iteration_count += 1
            state.current_step = "quality_correction"
            state.current_agent = AgentType.ORCHESTRATOR  # Let orchestrator decide correction strategy
            
            state.messages.append({
                "role": "assistant",
                "content": f"🔍 Quality check found {len(quality_result.issues_found)} issues (iteration {state.quality_iteration_count}). "
                          f"Overall quality: {quality_result.overall_quality_score:.1%}. Routing for corrections.",
                "agent": self.agent_type.value,
                "quality_score": quality_result.overall_quality_score,
                "issues_count": len(quality_result.issues_found)
            })
        else:
            # Quality acceptable or max iterations reached
            state.current_step = "completed"
            state.current_agent = AgentType.ORCHESTRATOR
            
            if quality_result.overall_quality_score >= 0.8:
                # Enhanced message based on validation type
                if quality_result.reference_form_used:
                    quality_message = (f"✅ Quality check passed! Overall quality: {quality_result.overall_quality_score:.1%} "
                                     f"({quality_result.passed_checks}/{quality_result.total_checks} checks passed)")
                else:
                    quality_message = (f"✅ Basic quality check passed! Overall quality: {quality_result.overall_quality_score:.1%} "
                                     f"({quality_result.passed_checks}/{quality_result.total_checks} basic checks passed) "
                                     f"⚠️ Note: Limited validation without reference form")
                
                state.messages.append({
                    "role": "assistant", 
                    "content": quality_message,
                    "agent": self.agent_type.value,
                    "quality_score": quality_result.overall_quality_score
                })
            else:
                state.messages.append({
                    "role": "assistant",
                    "content": f"⚠️  Quality check completed with issues. Quality: {quality_result.overall_quality_score:.1%} "
                              f"(max iterations reached: {state.quality_iteration_count})",
                    "agent": self.agent_type.value,
                    "quality_score": quality_result.overall_quality_score
                })
        
        return state
    
    def _handle_quality_error(self, state: AgentState, error_message: str) -> AgentState:
        """Handle quality checking errors."""
        print(f"❌ Quality Checker Error: {error_message}")
        
        state.messages.append({
            "role": "assistant",
            "content": f"❌ Quality checking failed: {error_message}",
            "agent": self.agent_type.value,
            "error": error_message
        })
        
        # Continue to completion despite error
        state.current_step = "completed"
        state.current_agent = AgentType.ORCHESTRATOR
        
        return state
    
    async def _extract_current_form_values(self, state: AgentState) -> Dict[str, Any]:
        """Extract current values from the filled form."""
        
        if not state.filled_form_path or not os.path.exists(state.filled_form_path):
            return {}
        
        file_extension = os.path.splitext(state.filled_form_path)[1].lower()
        
        if file_extension in ['.xlsx', '.xlsm']:
            return await self._extract_excel_values(state.filled_form_path)
        elif file_extension == '.pdf':
            return await self._extract_pdf_values(state.filled_form_path)
        else:
            return {}
    
    async def _extract_excel_values(self, form_path: str) -> Dict[str, Any]:
        """Extract values from Excel form."""
        
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(form_path, data_only=True)
            values = {}
            
            # Focus on first worksheet for now (can be made configurable)
            sheet = wb[wb.sheetnames[0]]
            
            # Extract values using the same pattern as reference analysis
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.column == 1 and cell.value and isinstance(cell.value, str):  # Labels in column A
                        field_label = str(cell.value).strip()
                        value_cell = sheet.cell(row=cell.row, column=2)  # Values in column B
                        
                        if value_cell.value is not None:
                            field_id = f"field_{cell.row-1}"  # Match the ID generation from reference analysis
                            values[field_id] = value_cell.value
            
            print(f"   📋 Extracted {len(values)} values from filled form")
            return values
            
        except Exception as e:
            print(f"   ❌ Error extracting Excel values: {str(e)}")
            return {}
    
    async def _extract_pdf_values(self, form_path: str) -> Dict[str, Any]:
        """Extract values from PDF form."""
        
        try:
            from src.tools.comprehensive_form_analyzer import ComprehensiveFormAnalysisTool
            
            # Use the comprehensive form analyzer to extract PDF form fields and values
            analyzer = ComprehensiveFormAnalysisTool()
            pdf_fields = analyzer._extract_pdf_form_fields(form_path)
            
            # Convert to the format expected by quality checker
            values = {}
            for field_name, field_data in pdf_fields.items():
                if field_data.get('current_value'):  # Only include fields with values
                    values[field_name] = field_data['current_value']
            
            print(f"   📄 Extracted {len(values)} filled values from PDF form")
            return values
            
        except Exception as e:
            print(f"   ❌ Error extracting PDF values: {str(e)}")
            return {}
        return {}
    
    def _check_semantic_consistency(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any, 
        all_current_values: Dict[str, Any],
        state: AgentState
    ) -> Optional[QualityIssue]:
        """Check if the current value semantically matches the expected pattern."""
        
        # Key semantic validation: Date field analysis
        if pattern.semantic_category == "document_date":
            return self._validate_document_date(pattern, current_value, all_current_values, state)
        elif pattern.semantic_category == "personal_date":
            return self._validate_personal_date(pattern, current_value, all_current_values)
        elif pattern.semantic_category == "personal_name":
            return self._validate_personal_name(pattern, current_value, all_current_values, state)
        elif pattern.semantic_category == "contact_info":
            return self._validate_contact_info(pattern, current_value)
        elif pattern.semantic_category == "numeric_value":
            return self._validate_numeric_value(pattern, current_value, all_current_values)
        
        return None
    
    def _validate_document_date(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any, 
        all_current_values: Dict[str, Any],
        state: AgentState
    ) -> Optional[QualityIssue]:
        """Validate document-related dates (like Eingangsdatum)."""
        
        current_str = str(current_value).strip()
        
        # Check if this looks like a personal birth date instead of document date
        birth_date_fields = [v for k, v in all_current_values.items() 
                           if any(birth_keyword in str(k).lower() or 
                                 (isinstance(v, str) and birth_keyword in v.lower())
                                 for birth_keyword in ['geburt', 'birth'])]
        
        # If the document date matches a birth date, it's likely wrong
        if any(str(birth_date) == current_str for birth_date in birth_date_fields):
            # Get reference example to suggest correct pattern
            reference_example = pattern.example_values[0] if pattern.example_values else "current date"
            
            return QualityIssue(
                field_id=pattern.field_id,
                field_name=pattern.field_name,
                issue_type="semantic_mismatch",
                current_value=current_value,
                expected_pattern="document_submission_date",
                confidence=0.85,
                suggestion=f"Field '{pattern.field_name}' contains a birth date but should contain a document date. "
                          f"Reference example: '{reference_example}'. Consider using submission/received date.",
                severity="high"
            )
        
        # Check if the date is in the future (might indicate wrong date)
        if self._looks_like_date(current_str):
            try:
                from datetime import datetime
                import re
                
                # Parse common date formats
                date_match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', current_str)
                if date_match:
                    day, month, year = map(int, date_match.groups())
                    parsed_date = datetime(year, month, day)
                    current_date = datetime.now()
                    
                    # If it's more than 1 year in the past, might be birth date
                    if (current_date - parsed_date).days > 365:
                        return QualityIssue(
                            field_id=pattern.field_id,
                            field_name=pattern.field_name,
                            issue_type="temporal_inconsistency",
                            current_value=current_value,
                            confidence=0.7,
                            suggestion=f"Document date '{current_value}' seems too old for a recent submission. "
                                      f"Verify this is the correct document received/submission date.",
                            severity="high"
                        )
                        
            except:
                pass  # If parsing fails, skip temporal validation
        
        return None
    
    def _validate_personal_date(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any, 
        all_current_values: Dict[str, Any]
    ) -> Optional[QualityIssue]:
        """Validate personal dates (like birth date)."""
        
        current_str = str(current_value).strip()
        
        # Check if birth date is realistic (not in future, not too old)
        if self._looks_like_date(current_str):
            try:
                from datetime import datetime
                import re
                
                date_match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', current_str)
                if date_match:
                    day, month, year = map(int, date_match.groups())
                    parsed_date = datetime(year, month, day)
                    current_date = datetime.now()
                    
                    age_years = (current_date - parsed_date).days / 365.25
                    
                    if age_years < 0:
                        return QualityIssue(
                            field_id=pattern.field_id,
                            field_name=pattern.field_name,
                            issue_type="temporal_inconsistency",
                            current_value=current_value,
                            confidence=0.95,
                            suggestion=f"Birth date cannot be in the future: '{current_value}'",
                            severity="critical"
                        )
                    elif age_years > 120:
                        return QualityIssue(
                            field_id=pattern.field_id,
                            field_name=pattern.field_name,
                            issue_type="temporal_inconsistency", 
                            current_value=current_value,
                            confidence=0.9,
                            suggestion=f"Birth date '{current_value}' indicates unrealistic age of {age_years:.0f} years",
                            severity="high"
                        )
                        
            except:
                pass
        
        return None
    
    def _validate_personal_name(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any, 
        all_current_values: Dict[str, Any],
        state: AgentState
    ) -> Optional[QualityIssue]:
        """Validate personal names for consistency."""
        
        current_str = str(current_value).strip()
        
        # Check for common name issues
        if len(current_str) < 2:
            return QualityIssue(
                field_id=pattern.field_id,
                field_name=pattern.field_name,
                issue_type="format_error",
                current_value=current_value,
                confidence=0.8,
                suggestion=f"Name '{current_value}' seems too short. Verify this is complete.",
                severity="medium"
            )
        
        # Check if name contains numbers (usually not valid)
        if any(char.isdigit() for char in current_str):
            return QualityIssue(
                field_id=pattern.field_id,
                field_name=pattern.field_name,
                issue_type="format_error",
                current_value=current_value,
                confidence=0.9,
                suggestion=f"Name '{current_value}' contains numbers. Verify this is correct.",
                severity="medium"
            )
        
        return None
    
    def _validate_contact_info(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any
    ) -> Optional[QualityIssue]:
        """Validate contact information."""
        
        current_str = str(current_value).strip()
        
        # Email validation
        if pattern.field_type == "email" or '@' in current_str:
            if '@' not in current_str or '.' not in current_str.split('@')[-1]:
                return QualityIssue(
                    field_id=pattern.field_id,
                    field_name=pattern.field_name,
                    issue_type="format_error",
                    current_value=current_value,
                    confidence=0.9,
                    suggestion=f"Email format appears invalid: '{current_value}'",
                    severity="high"
                )
        
        return None
    
    def _validate_numeric_value(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any, 
        all_current_values: Dict[str, Any]
    ) -> Optional[QualityIssue]:
        """Validate numeric values for reasonableness."""
        
        try:
            numeric_value = float(current_value)
            
            # Age validation
            if 'alter' in pattern.field_name.lower() or 'age' in pattern.field_name.lower():
                if numeric_value < 0 or numeric_value > 120:
                    return QualityIssue(
                        field_id=pattern.field_id,
                        field_name=pattern.field_name,
                        issue_type="range_error",
                        current_value=current_value,
                        confidence=0.9,
                        suggestion=f"Age value '{current_value}' seems unrealistic",
                        severity="high"
                    )
            
            # Grade validation (German system typically 1-6)
            if 'note' in pattern.field_name.lower() or 'grade' in pattern.field_name.lower():
                if numeric_value < 1 or numeric_value > 6:
                    return QualityIssue(
                        field_id=pattern.field_id,
                        field_name=pattern.field_name,
                        issue_type="range_error",
                        current_value=current_value,
                        confidence=0.8,
                        suggestion=f"Grade '{current_value}' outside typical range (1-6)",
                        severity="medium"
                    )
                        
        except (ValueError, TypeError):
            # If it should be numeric but isn't
            if pattern.field_type == "number":
                return QualityIssue(
                    field_id=pattern.field_id,
                    field_name=pattern.field_name,
                    issue_type="data_type_error",
                    current_value=current_value,
                    confidence=0.9,
                    suggestion=f"Expected numeric value but got: '{current_value}'",
                    severity="medium"
                )
        
        return None
    
    def _check_format_consistency(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any
    ) -> Optional[QualityIssue]:
        """Check if the format matches the expected pattern."""
        
        if pattern.value_pattern == "date_format" and not self._looks_like_date(str(current_value)):
            return QualityIssue(
                field_id=pattern.field_id,
                field_name=pattern.field_name,
                issue_type="format_error",
                current_value=current_value,
                expected_pattern=pattern.value_pattern,
                confidence=0.7,
                suggestion=f"Expected date format but got: '{current_value}'",
                severity="medium"
            )
        
        return None
    
    async def _check_contextual_consistency(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any, 
        all_current_values: Dict[str, Any],
        all_patterns: Dict[str, ReferenceFieldPattern],
        state: AgentState
    ) -> Optional[QualityIssue]:
        """Check contextual consistency using all available knowledge."""
        
        # Cross-field validation using extracted document data
        if state.extracted_data:
            contextual_issue = await self._validate_against_source_documents(
                pattern, current_value, state.extracted_data, state
            )
            if contextual_issue:
                return contextual_issue
        
        # Cross-reference with other fields in the form
        cross_ref_issue = self._validate_cross_field_consistency(
            pattern, current_value, all_current_values, all_patterns
        )
        if cross_ref_issue:
            return cross_ref_issue
        
        return None
    
    async def _validate_against_source_documents(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any, 
        extracted_data: Dict[str, Any],
        state: AgentState
    ) -> Optional[QualityIssue]:
        """Validate against original source documents for deeper context."""
        
        # Use LLM to assess if the value makes sense given the source documents
        if pattern.semantic_category == "document_date":
            # For document dates, check if there are better alternatives in the source
            prompt = f"""
            Field Analysis Task:
            
            Field: {pattern.field_name} (ID: {pattern.field_id})
            Current Value: {current_value}
            Field Purpose: Document submission/received date
            
            Available Source Data: {json.dumps(extracted_data, indent=2)}
            
            Question: Is the current value appropriate for a document submission/received date?
            
            Consider:
            1. Is this value a personal date (birth date) instead of document date?
            2. Are there better date options in the source data?
            3. Should this be today's date or a recent date?
            
            Respond with JSON:
            {{
                "is_appropriate": true/false,
                "confidence": 0.0-1.0,
                "issue_description": "brief description if not appropriate",
                "suggested_alternative": "better value if available"
            }}
            """
            
            try:
                response = await self.llm_client.generate_response(prompt)
                # Parse LLM response (simplified - would need proper JSON parsing)
                if "false" in response.lower() and ("birth" in response.lower() or "personal" in response.lower()):
                    return QualityIssue(
                        field_id=pattern.field_id,
                        field_name=pattern.field_name,
                        issue_type="contextual_error",
                        current_value=current_value,
                        confidence=0.8,
                        suggestion=f"LLM analysis suggests '{pattern.field_name}' contains personal data instead of document date. "
                                  f"Consider using current date or document submission date.",
                        severity="high"
                    )
            except:
                pass  # If LLM call fails, skip this validation
        
        return None
    
    def _validate_cross_field_consistency(
        self, 
        pattern: ReferenceFieldPattern, 
        current_value: Any,
        all_current_values: Dict[str, Any],
        all_patterns: Dict[str, ReferenceFieldPattern]
    ) -> Optional[QualityIssue]:
        """Validate consistency across multiple fields."""
        
        # Age vs birth date consistency
        if pattern.semantic_category == "numeric_value" and "alter" in pattern.field_name.lower():
            # Find birth date field
            birth_date_value = None
            for field_id, other_pattern in all_patterns.items():
                if other_pattern.semantic_category == "personal_date" and field_id in all_current_values:
                    birth_date_value = all_current_values[field_id]
                    break
            
            if birth_date_value and self._looks_like_date(str(birth_date_value)):
                try:
                    from datetime import datetime
                    import re
                    
                    # Calculate expected age from birth date
                    date_match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', str(birth_date_value))
                    if date_match:
                        day, month, year = map(int, date_match.groups())
                        birth_date = datetime(year, month, day)
                        expected_age = int((datetime.now() - birth_date).days / 365.25)
                        
                        current_age = int(float(current_value))
                        
                        if abs(current_age - expected_age) > 1:  # Allow 1 year difference
                            return QualityIssue(
                                field_id=pattern.field_id,
                                field_name=pattern.field_name,
                                issue_type="consistency_error",
                                current_value=current_value,
                                confidence=0.9,
                                suggestion=f"Age '{current_value}' inconsistent with birth date '{birth_date_value}'. "
                                          f"Expected age: {expected_age}",
                                severity="high"
                            )
                except:
                    pass
        
        return None
    
    async def _save_quality_report(self, quality_result: QualityAssessmentResult, state: AgentState):
        """Save detailed quality report for debugging."""
        try:
            import json
            from datetime import datetime
            from ..config import config
            
            # Create report data
            report_data = {
                "timestamp": quality_result.assessment_timestamp,
                "overall_score": quality_result.overall_quality_score,
                "passed_checks": quality_result.passed_checks,
                "total_checks": quality_result.total_checks,
                "reference_form_used": quality_result.reference_form_used,
                "requires_correction": quality_result.requires_correction,
                "issues": [
                    {
                        "field_id": issue.field_id,
                        "field_name": issue.field_name,
                        "issue_type": issue.issue_type,
                        "current_value": str(issue.current_value) if issue.current_value is not None else None,
                        "expected_pattern": issue.expected_pattern,
                        "confidence": issue.confidence,
                        "suggestion": issue.suggestion,
                        "severity": issue.severity
                    }
                    for issue in quality_result.issues_found
                ],
                "quality_iteration": state.quality_iteration_count
            }
            
            # Save to output directory
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"{config.OUTPUT_DIR}/quality_assessment_{timestamp}.json"
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)
                
            print(f"   📊 Quality report saved: {output_path}")
            
        except Exception as e:
            print(f"   ⚠️  Could not save quality report: {str(e)}")

    def _truncate_value(self, value: Any, max_length: int = 50) -> str:
        """Truncate value for display."""
        value_str = str(value)
        return value_str[:max_length] + "..." if len(value_str) > max_length else value_str